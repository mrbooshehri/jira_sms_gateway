import os
from flask import Flask, request, jsonify
from kavenegar import *
from openpyxl import load_workbook

app = Flask(__name__)

phonebook = {}

def read_excel_file_serveral_sheets(file_path):
    """
    Reads an Excel file with multiple sheets and returns a list of dictionaries.
    Each dictionary contains data from a row in the sheet, with an additional 'group' field.
    """
    wb = load_workbook(file_path)
    all_data = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Assuming the first column contains names and the second column contains numbers
            name, number = row
            number = str(number)
            if not number.startswith('0'):
                number = '0' + number
            # Create a dictionary for the current row, including the 'group' field
            entry = {
                'group': sheet_name,
                'name': name,
                'number': number
            }
            all_data.append(entry)

    return all_data

def read_excel_file(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active
    phonebook_local = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, number = row
        number = str(number)
        if not number.startswith('0'):
            number = '0' + number
        phonebook_local[name] = number
    return phonebook_local


def read_file(file_path):

    # Read receptor numbers from a file
    receptor_file_path = os.environ.get(os.environ['RECEPTOR_FILE_PATH'], 'receptors.txt')
    with open(receptor_file_path, 'r') as file:
        receptor_numbers = file.readlines()
    return receptor_numbers

# Load the phonebook at startup
excel_file_path = os.environ.get('EXCEL_FILE_PATH', 'phonebook.xlsx')
phonebook = read_excel_file(excel_file_path)

@app.route('/webhook', methods=['POST'])
def handle_webhook():
    # Extract data from the webhook payload
    data = request.json
    issue_user = data['user']['displayName']
    issue_key = data['issue']['key']
    issue_url = "https://pmo.vaslapp.com/browse/{}".format(issue_key)
    issue_priority = data['issue']['fields']['priority']['name']
    issue_assignee = data['issue']['fields']['assignee']['displayName']
    issue_creator= data['issue']['fields']['creator']['displayName']
    issue_reporter= data['issue']['fields']['reporter']['displayName']
    issue_type= data['issue']['fields']['issuetype']['name']
    issue_project_name= data['issue']['fields']['project']['name']
    issue_summary= data['issue']['fields']['summary']

    # Prepare the SMS message
    message = "User: {}\nKey: {}\nPriority: {}\nAssignee: {}\nCreator: {}\nReporter: {}\nType: {}\nProject Name: {}\nSummary: {}".format(issue_user, issue_key,issue_priority, issue_assignee, issue_creator, issue_reporter, issue_type, issue_project_name, issue_summary)

    # Read receptor numbers from a file
    #receptor_file_path = os.environ.get(os.environ['RECEPTOR_FILE_PATH'], 'receptors.txt')
    #with open(receptor_file_path, 'r') as file:
    #    receptor_numbers = file.readlines()

    receptor_numbers = phonebook.values()


#    # Check if the issue_user exists in the phonebook
#    if issue_user in phonebook:
#        # Retrieve the phone number associated with the issue_user
#        receptor_number = phonebook[issue_user]
#
#        # Send SMS using Kavenegar API
#        try:
#            api = KavenegarAPI(os.environ['KAVENEGAR_API_KEY'])
#            params = {
#                'receptor': receptor_number,
#                'message': message
#            }
#            response = api.sms_send(params)
#            print(f"SMS sent to {receptor_number}: {str(response)}")
#            return jsonify({"message": "SMS sent successfully"}), 200
#        except APIException as e:
#            print(str(e))
#            return jsonify({"error": "Failed to send SMS", "details": str(e)}), 500
#        except HTTPException as e:
#            print(str(e))
#            return jsonify({"error": "Failed to send SMS", "details": str(e)}), 500
#    else:
#        print(f"No phone number found for user: {issue_user}")
#        return jsonify({"error": "No phone number found for user"}), 404





    # Send SMS using Kavenegar API
    try:
        api = KavenegarAPI(os.environ['KAVENEGAR_API_KEY'])
        for number in receptor_numbers:
            number = number.strip() # Remove any leading/trailing whitespace
            params = {
                'receptor': number,
                'message': message
            }
            response = api.sms_send(params)
            print(f"SMS sent to {number}: {str(response)}")
        return jsonify({"message": "SMS sent successfully to all receptors"}), 200
    except APIException as e:
        print(str(e))
        return jsonify({"error": "Failed to send SMS", "details": str(e)}), 500
    except HTTPException as e:
        print(str(e))
        return jsonify({"error": "Failed to send SMS", "details": str(e)}), 500
    except FileNotFoundError:
        print(f"Receptor file not found: {receptor_file_path}")
        return jsonify({"error": "Receptor file not found"}), 500

if __name__ == '__main__':
    app.run(port=5000)
